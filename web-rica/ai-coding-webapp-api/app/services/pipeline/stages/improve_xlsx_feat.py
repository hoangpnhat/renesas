from pathlib import Path
from openpyxl import load_workbook
from spire.xls import Workbook
import os
from loguru import logger

# ------------- Convert and fix Excel files using LibreOffice in headless mode. -------------
def fix_excel_with_libreoffice(input_path, output_dir):
    """
    Convert and fix Excel files using LibreOffice in headless mode.
    
    This function uses LibreOffice's command-line interface to convert
    Excel files, which can help fix formatting issues and ensure compatibility.
    
    Args:
        input_path (str): Path to the input Excel file to be fixed
        output_dir (str): Directory path where the converted file will be saved
    
    Raises:
        subprocess.CalledProcessError: If LibreOffice conversion fails
    
    Example:
        >>> fix_excel_with_libreoffice(
        ...     input_path='corrupted.xlsx',
        ...     output_dir='./fixed'
        ... )
        File fixed and saved to ./fixed
    
    Note:
        On Windows, you may need to use the full path to soffice.exe:
        'C:\\Program Files\\LibreOffice\\program\\soffice.exe'
    """
    import subprocess
    # Command to run LibreOffice headless
    command = [
        'soffice',
        '--headless',
        '--convert-to', 'xlsx',
        '--outdir', output_dir,
        input_path
    ]
    
    try:
        subprocess.run(command, check=True, capture_output=True)
        logger.info(f"File fixed and saved to {output_dir}")
    except subprocess.CalledProcessError as e:
        logger.error(f"LibreOffice conversion failed for {input_path}: {e}", exc_info=True)
        raise

# ------------- Optimize Excel sheets by removing empty rows/columns and setting print areas -------------

def optimize_sheet_for_render(input_file, output_file, sheet_names):
    """
    Optimize Excel sheets by removing empty rows/columns and setting print areas.
    
    This function processes specified sheets in an Excel workbook to:
    - Remove trailing empty rows and columns
    - Set print areas to actual data boundaries
    - Fix rendering issues for LibreOffice compatibility
    
    Args:
        input_file (str): Path to the input Excel file
        output_file (str): Path to save the optimized Excel file
        sheet_names (list): List of sheet names to optimize
    
    Returns:
        None
    
    Example:
        >>> optimize_sheet_for_render(
        ...     input_file='data.xlsx',
        ...     output_file='optimized.xlsx',
        ...     sheet_names=['Sheet1', 'Sheet2']
        ... )
    """
    from spire.xls import Workbook
    workbook = Workbook()
    workbook.LoadFromFile(input_file)
    
    for sheet_name in sheet_names:
        sheet = workbook.Worksheets[sheet_name]
        
        # 1. Get the maximum boundaries
        data_range = sheet.AllocatedRange
        
        # Safety check: If the sheet is completely empty, skip it
        if data_range is None:
            logger.warning(f"Sheet '{sheet_name}' is completely empty. Skipping.")
            continue
        
        last_row = data_range.LastRow
        last_col = data_range.LastColumn
        #print(f"Original boundary for '{sheet_name}': Row {last_row}, Col {last_col}")
        
        # 2. Loop BACKWARDS to safely delete empty rows
        for r in range(last_row, 0, -1):
            row_obj = sheet.Rows[r - 1]
            if row_obj is None or row_obj.IsBlank:
                sheet.DeleteRow(r)
        
        # 3. Loop BACKWARDS to safely delete empty columns
        for c in range(last_col, 0, -1):
            col_obj = sheet.Columns[c - 1]
            if col_obj is None or col_obj.IsBlank:
                sheet.DeleteColumn(c)
        
        # 4. Re-calculate the new shrunken boundaries
        new_range = sheet.AllocatedRange
        
        # 5. CRITICAL FIX FOR LIBREOFFICE: Set the Print Area
        if new_range is not None:
            sheet.PageSetup.PrintArea = new_range.RangeAddressLocal
            #print(f"Optimized boundary for '{sheet_name}': {new_range.RangeAddressLocal}")
    
    workbook.SaveToFile(output_file)
    workbook.Dispose()
    #print("Optimization complete.")

# ------------- Remove images, fix formatting, and optimize Excel file. -------------
def clean_and_optimize_excel(excel_path, output_dir):
    """Remove images, fix formatting, and optimize Excel file."""
    
    # get excel name:
    xlsx_name = Path(excel_path).stem
    # Remove images and drawings
    wb = load_workbook(excel_path)
    sheet_names = wb.sheetnames
    for ws in wb.worksheets:
        ws._images = []
        if hasattr(ws, 'drawing') and ws.drawing:
            ws.drawing = None
    temp_dir = Path(output_dir,'optimize',xlsx_name, "temp")
    temp_dir.mkdir(exist_ok=True,parents=True)
    remove_img_shape_path = Path(temp_dir,'remv_img_shape.xlsx')
    wb.save(str(remove_img_shape_path))
    logger.info(">>> Step 1: Remove all images, Shapes and OLE objects - DONE")
    
    # Fix with LibreOffice
    clean_dir = Path(output_dir,'optimize',xlsx_name, "clean")
    clean_dir.mkdir(exist_ok=True,parents=True)
    fix_excel_with_libreoffice(input_path=remove_img_shape_path, output_dir=str(clean_dir))
    
    
    # Fix format with Spire.XLS
    workbook = Workbook()
    workbook.LoadFromFile(f"{clean_dir}/remv_img_shape.xlsx")
    fix_clean_path = f'{clean_dir}/fix_cleaned.xlsx'
    workbook.SaveToFile(fix_clean_path)
    workbook.Dispose()
    logger.info(">>> Step 2: Fix File excel is created by Openpyxl - DONE")

    # Optimize sheet
    optimize_excel_path = f'{clean_dir}/optimized_excel.xlsx'
    optimize_sheet_for_render(input_file=fix_clean_path,
                              output_file=optimize_excel_path,
                              sheet_names=sheet_names)
    
    logger.info(">>> Step 3: Remove all empty cols and rows - DONE")
    return optimize_excel_path


# -------------------  Analyze Excel sheet dimensions and cell info, then adjust with Spire. -------------------
# Parse range of sheet:
def get_sheet_info(range_string):
    """
    Parse an Excel-style range string (e.g., 'A1:P71') into a dictionary.
    
    Args:
        range_string: String in format 'A1:P71'
        
    Returns:
        Dictionary with start/end column names, column IDs, and row IDs
    """
    # Standard Excel defaults if no custom dimension is set
    DEFAULT_COL_WIDTH = 8.43
    DEFAULT_ROW_HEIGHT = 15.0

    # Split the range into start and end cells
    start_cell, end_cell = range_string.split(':')
    
    # Helper function to separate column letters from row numbers
    def parse_cell(cell):
        col_name = ''
        row_num = ''
        for char in cell:
            if char.isalpha():
                col_name += char
            else:
                row_num += char
        return col_name, int(row_num)
    
    # Helper function to convert column name to ID (A=1, B=2, ..., Z=26, AA=27, etc.)
    def col_name_to_id(col_name):
        col_id = 0
        for char in col_name:
            col_id = col_id * 26 + (ord(char.upper()) - ord('A') + 1)
        return col_id
    
    # Parse start and end cells
    start_col_name, start_row_id = parse_cell(start_cell)
    end_col_name, end_row_id = parse_cell(end_cell)
    
    # Convert column names to IDs
    start_col_id = col_name_to_id(start_col_name)
    end_col_id = col_name_to_id(end_col_name)

    # range col, width
    range_col = len(range(start_col_id, end_col_id + 1))
    range_row = len(range(start_row_id, end_row_id + 1))

    # default excel row height is 15, default excel column width is 8.43
    total_default_width = range_col * DEFAULT_COL_WIDTH
    total_default_height = range_row * DEFAULT_ROW_HEIGHT

    return {
        'start_col_name': start_col_name,
        'start_col_id': start_col_id,
        'start_row_id': start_row_id,
        'end_col_name': end_col_name,
        'end_col_id': end_col_id,
        'end_row_id': end_row_id,
        'range_col': range_col,
        'range_row': range_row,
        'total_default_width': total_default_width,
        'total_default_height': total_default_height
    }

# Get Height
def calculate_wrap_height(text: str, default_height: float = 20.0, char_limit_per_line: int = 50) -> float:
    """
    Calculates cell height by multiplying default height by the total number of lines.
    Accounts for both explicit newlines and physical text wrapping.
    """
    import math
    if not text:
        return default_height

    # 1. Split text by explicit hard returns
    explicit_lines = text.split('\n')
    total_lines = 0

    # 2. Calculate the total lines needed after physical wrapping
    for line in explicit_lines:
        line_length = len(line)
        
        if line_length > char_limit_per_line:
            # Ceiling division: e.g., 55 chars / 50 limit = 1.1 -> 2 lines
            total_lines += math.ceil(line_length / char_limit_per_line)
        else:
            # Fits on a single line
            total_lines += 1

    if total_lines == 1:
        return default_height
    else:
        return total_lines * default_height*0.6  # Adjust multiplier as needed for better fit

# Get all Cell Info for Autofit
from openpyxl.worksheet.worksheet import Worksheet
def get_all_cell_info(source_ws: Worksheet) ->list:
    """Get info of all cell in the worksheet (if having values) for autofit by openpyxl
    Args: 
        source_ws: openpyxl worksheet object
    Returns: 
        list: List of dict with column letter, index and max length
    """
    col_data = []
    for idx, col in enumerate(source_ws.columns):
        for cell in col:
            if cell.value:
                
                # Calculate height based on content and wrapping
                cell_str = str(cell.value)
                cell_height = calculate_wrap_height(cell_str)

                # Heuristic for width: length of the longest line
                lines = cell_str.split('\n')
                max_len: float = 0
                for line in lines:
                    if len(line) > max_len:
                        # calculate the font size coeff
                        f_size = cell.font.sz if cell.font and cell.font.sz else 10
                        if f_size > 10:
                            size_coeff = (f_size)/10 * 0.9
                        else:
                            size_coeff = 1
                        max_len = len(line) * size_coeff
                
                # Check is Wrap Text and Merged Cell 
                is_wrap_text = True if (cell.alignment and cell.alignment.wrap_text) else False
                is_merged_cell = cell.coordinate in source_ws.merged_cells 

                # Append cell info to list 
                col_data.append({'column': cell.column_letter, 'col_id': cell.column, 'row_id': cell.row,
                                 'is_wrap_text': is_wrap_text, 'is_merged_cell': is_merged_cell,
                                 'width': max_len, 'height': cell_height,'text':cell_str,'size':cell.font.sz}) #'cordinate': cell.coordinate,
                
    # Return data for all cells having values in the sheet
    return col_data

# Calculate optimal row heights and column widths using cell metrics and positional coefficients.
def summarize_sheet_dimensions(input_data: dict) -> dict:
    """
    Calculate optimal row heights and column widths using cell metrics and positional coefficients.
    
    Args:
        input_data (dict): Contains 'cell_info_ls', 'total_default_width', 'end_col_id', 'range_col', 'sheet_name'
    
    Returns:
        dict: Sheet summary with 'sheet_name', 'column_ls', 'row_ls' containing adjusted dimensions
    
    Example:
        result = summarize_sheet_dimensions({'cell_info_ls': cells, 'sheet_name': 'Sheet1', ...})
    """
    cell_info_ls = input_data.get('cell_info_ls', [])
    total_default_width = input_data.get('total_default_width', 1.0)
    
    # Handle edge case to prevent division by zero
    if total_default_width <= 0:
        total_default_width = 8.43 # the default width of one row in Excel
        
    end_col_id = input_data.get('end_col_id', 0)

    # --- 1. Process Row Heights ---
    row_dict = {}
    for cell in cell_info_ls:
        r_id = cell.get('row_id')
        height = cell.get('height', 15.0)
        
        # Get the tallest height for each row
        if r_id not in row_dict:
            row_dict[r_id] = height
        else:
            row_dict[r_id] = max(row_dict[r_id], height)
            
    row_ls = [{'row_id': r, 'height': h} for r, h in sorted(row_dict.items())]
    #print(f"Row Heights: {row_ls}")

    # --- 2. Process Column Widths ---
    col_group = {}
    for cell in cell_info_ls:
        c_id = cell.get('col_id')
        c_name = cell.get('column')
        
        if c_id not in col_group:
            col_group[c_id] = {'name': c_name, 'widths': []} #'lengths': [], 
        
        #col_group[c_id]['lengths'].append(text_len)
        col_group[c_id]['widths'].append(cell.get('width', 8.0))

    #print(f"Column Grouping: {col_group}")

    column_ls = []
    for c_id, data in sorted(col_group.items()):
        max_width = max(data['widths']) if data['widths'] else 0
        avg_width = sum(data['widths']) / len(data['widths']) if data['widths'] else 8.0
        
        # Calculate Width coefficients and Close_coeff
        width_coeff = max_width / total_default_width
        
        # Assuming input col_id in cell_info_ls start from 1 (e.g., A=1), 
        close_coeff = end_col_id - c_id + 1

        # Apply Rule Set
        col_range = input_data.get('range_col')
        coeff = width_coeff * (col_range/close_coeff) # combine coeff helps adjust the excel width more balance
        if close_coeff == 1 or coeff <= 1:
            # Rule 3: Specially, in the last columns, set the max length
            if max_width < 50 or close_coeff == 1:
                final_width = round(max_width,2)
            else:
                final_width = round(max_width*(1-close_coeff/col_range),2)
            #print(f"Col {data['name']} (ID {c_id}): adjust_coeff={coeff} <= 1, setting final_width ={final_width} (max_width={max_width})")
        else:
            # Rule 2: Based on 2 coeff, generate suitable rule using average width
            # Logic: Base = average. 
            # Add bonus for high width_coeff. Subtract penalty if far from end (high close_coeff).
            bonus_width = (max_width * 0.5)
            position_penalty = (close_coeff* 8 * 0.6)
            calculated_width = avg_width + bonus_width - position_penalty
            
            # Constrain the width so it makes logical sense (between average and max)
            final_width = max(avg_width, min(calculated_width, max_width))
            #print(f"Col {data['name']} (ID {c_id}): max_width={max_width}, avg_width={avg_width:.2f}, width_coeff={width_coeff:.2f}, close_coeff={close_coeff}, calculated_width={calculated_width:.2f} -> final_width={final_width:.2f}")

        column_ls.append({
            'col_name': data['name'],
            'col_id': c_id,
            'width': round(final_width, 2)
        })

    # --- 3. Return Summary Dictionary ---
    return {
        'sheet_name': input_data['sheet_name'],
        'column_ls': column_ls,
        'row_ls': row_ls
    }

# Applies calculated column widths and row heights to an Excel workbook using Spire.XLS.
def adjust_dimensions_with_spire(file_path: str, output_path: str, full_sheet_ls: list):
    """
    Applies calculated column widths and row heights to an Excel workbook using Spire.XLS.
    
    This function adjusts dimensions only when the target size is larger than the current size,
    ensuring content is never truncated. All dimensions are capped at Excel's maximum limits.
    
    Args:
        file_path: Path to the source Excel file to be processed.
        output_path: Path where the modified Excel file will be saved.
        full_sheet_ls: List of dictionaries containing dimension data for each sheet.
            Each dictionary should have:
            - "sheet_name" (str): Name of the worksheet
            - "column_ls" (list): List of dicts with "col_id" (0-indexed int), 
                                  "col_name" (str), and "width" (float)
            - "row_ls" (list): List of dicts with "row_id" (1-indexed int) 
                               and "height" (float)
    
    Behavior:
        - Column widths are capped at 255 (Excel's maximum)
        - Row heights are capped at 409.5 (Excel's maximum)
        - Dimensions are only increased, never decreased (preserves existing sizing)
        - Only positive target dimensions are applied
    
    Returns:
        None. The modified workbook is saved to output_path.
    """
    workbook = Workbook()
    workbook.LoadFromFile(file_path)
    
    # Iterate through each sheet's data in the provided list
    for sheet_data in full_sheet_ls:

        sheet_name = sheet_data.get("sheet_name")
        logger.info(f"\t - Adjust the Sheet: {sheet_name}")
        column_ls = sheet_data.get("column_ls", [])
        row_ls = sheet_data.get("row_ls", [])
        
        # Access the specific sheet
        sheet = workbook.Worksheets[sheet_name]
        
        # 1. Apply Column Widths
        for col_info in column_ls:
            # Spire columns are 1-indexed; assuming col_id is 0-indexed
            spire_col_idx = col_info["col_id"] #+ 1 
            target_width = col_info["width"]
            
            if target_width > 0:
                # Cap width at 255 to avoid exceeding Excel's maximum column width limit
                current_width = sheet.GetColumnWidth(spire_col_idx)
                if current_width >= target_width:
                    final_width = current_width
                else:
                    final_width = min(target_width, 255)
                
                #print(f"Setting width for column {col_info['col_name']} (Spire idx {spire_col_idx}): Target={target_width}, Final={final_width}, Current={current_width}")
                sheet.SetColumnWidth(spire_col_idx, final_width)
                
        # 2. Apply Row Heights
        for row_info in row_ls:
            # Spire rows are 1-indexed; assuming row_id from input is already 1-indexed
            spire_row_idx = row_info["row_id"]
            target_height = row_info["height"]
            
            if target_height > 0:
                # Cap height at 409.5 to avoid exceeding Excel's maximum row height limit
                current_height = sheet.GetRowHeight(spire_row_idx)
                if current_height >= target_height:
                    final_height = current_height
                else:
                    final_height = min(target_height, 409.5)
                    
                #print(f"Setting height for row {spire_row_idx}: Target={target_height}, Final={final_height}, Current={current_height}")
                sheet.SetRowHeight(spire_row_idx, final_height)

    workbook.SaveToFile(output_path)
    workbook.Dispose()
    

# Full Flow for Enhacing Excel Feature 
def analyze_and_adjust_excel_dimensions(file_path, output_dir):
    """
    Analyze Excel sheet dimensions and cell info, then adjust with Spire.
    
    Args:
        file_path (str): Path to Excel file to analyze
        output_dir (str): Directory for output file
    
    Returns:
        list: Sheet adjustment information for each sheet
    
    Example:
        adjust_data = analyze_and_adjust_excel_dimensions('data.xlsx', './output')
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path, data_only=True)
    adjust_excel_ls = []
    for sheet_name in wb.sheetnames:
        #print(f"Sheet Name: {sheet_name}")
        sheet = wb[sheet_name]
        
        # Get sheet dimensions and cell info
        dimensions = sheet.dimensions
        if not dimensions or ':' not in dimensions:
            logger.warning(f"Sheet '{sheet_name}' has no valid dimensions. Skipping.")
            continue
        sheet_info = get_sheet_info(dimensions)
        all_cell_info_ls = get_all_cell_info(source_ws=sheet)
        
        sheet_info['sheet_name'] = sheet_name
        sheet_info['cell_info_ls'] = all_cell_info_ls
        
        # Summarize and append
        adj_info_sheet = summarize_sheet_dimensions(input_data=sheet_info)
        adjust_excel_ls.append(adj_info_sheet)
    
    # Adjust dimensions with Spire
    output_path = f'{output_dir}/final_adjusted.xlsx'
    adjust_dimensions_with_spire(file_path=file_path, output_path=output_path, 
                                 full_sheet_ls=adjust_excel_ls)  
    logger.info(f">>> Final Step: Complete to save Adjusted excel file in {output_path}")
    
    return output_path, adjust_excel_ls


# ---- CHECK SHEET IMAGE SIZE ----
# --- Get one Image Size ---
def get_image_dimensions(image_path):
    """Get width and height of an image without fully loading it into memory"""
    from PIL import Image
    try:
        with Image.open(image_path) as img:
            width, height = img.size
            return width, height
    except Exception as e:
        logger.error(f"Error reading {image_path}: {e}")
        return None, None

# ---- CHECK ALL SHEET IMAGE ----
def check_all_sheet_images(image_directory, max_height=5000, max_width=3000):
    """Check dimensions of all images in directory"""
    import re 
    results = []

    for filename in os.listdir(image_directory):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
            image_path = os.path.join(image_directory, filename)
            width, height = get_image_dimensions(image_path)

            # Get page
            _str = Path(filename).stem
            pat = re.compile(r'page_[0-9]+$')
            match = pat.search(_str)
            if match is None:
                logger.warning(f"Filename '{filename}' does not match expected page pattern. Skipping.")
                continue
            page = match.group()
            
            if width and height:
                is_too_long = height > max_height
                is_too_wide = width > max_width
                
                results.append({
                    'filename': image_path,
                    'page': page,
                    'width': width,
                    'height': height,
                    'too_long': is_too_long,
                    'too_wide': is_too_wide,
                    'use_method_2': is_too_long or is_too_wide
                })
    
    return results