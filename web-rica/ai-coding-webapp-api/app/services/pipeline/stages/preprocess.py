import os
import tempfile
import numpy as np
import openpyxl
from abc import ABC, abstractmethod
from pathlib import Path
from loguru import logger

from ..state.tracker import PageStateTracker

# Import excel feature
from .improve_xlsx_feat import (check_all_sheet_images,clean_and_optimize_excel, analyze_and_adjust_excel_dimensions)


class BasePreprocessStage(ABC):
    def __init__(self, page_state_tracker: PageStateTracker | None = None):
        self.page_state_tracker = page_state_tracker

    @abstractmethod
    def preprocess(self, file_bytes: bytes) -> bytes:
        pass


class ExcelPreprocessStage(BasePreprocessStage):

    def __init__(self, page_state_tracker: PageStateTracker | None = None):
        super().__init__(page_state_tracker)
        self.sheet_dict: dict = {}
        self.sheet_names: list = []

    def get_max_length_all_cols(self, source_ws) -> list:
        """Get max byte-length of all columns in a worksheet for openpyxl autofit.

        Args:
            source_ws: openpyxl worksheet object.

        Returns:
            list: One dict per column with keys 'column', 'index', 'max_length'.
        """
        col_data = []
        for idx, col in enumerate(source_ws.columns):
            column_letter = col[0].column_letter if hasattr(col[0], 'column_letter') else None
            if not column_letter:
                continue

            max_length = 0
            for cell in col:
                # Skip merged cells without column_letter attribute
                if not hasattr(cell, 'column_letter'):
                    continue
                if cell.value:
                    # Byte length gives better scale for CJK characters
                    length = len(str(cell.value).encode('utf-8'))
                    max_length = max(max_length, length)
                    col_data.append({'column': column_letter, 'index': idx, 'max_length': max_length})

        return col_data

    def get_last_column_autofit(self, excel_path: str, sheet_name: str | None = None, get_all_sheets: bool = False) -> dict:
        """Compute last-column autofit width info for an Excel file.

        When ``get_all_sheets=True``, results are also stored on ``self.sheet_dict``
        and ``self.sheet_names`` for use by subsequent processing steps.

        Args:
            excel_path (str): Path to the Excel file.
            sheet_name (str, optional): Sheet name to process. Required when
                ``get_all_sheets`` is False.
            get_all_sheets (bool): If True, process all sheets. Defaults to False.

        Returns:
            dict: When ``get_all_sheets=True``, maps each sheet name to its
                last-column info dict. When ``get_all_sheets=False``, maps the
                sheet name to a list of per-column dicts.

        Raises:
            KeyError: If ``sheet_name`` is not found in the workbook.
        """
        source_wb = openpyxl.load_workbook(excel_path, data_only=True)
        try:
            sheet_ls = source_wb.sheetnames

            if get_all_sheets:
                sheet_dict = {}
                for name in sheet_ls:
                    logger.info(f"Processing sheet: {name}")
                    source_ws = source_wb[name]
                    col_ls = self.get_max_length_all_cols(source_ws)
                    if not col_ls:
                        continue
                    avg_length = np.mean([c['max_length'] for c in col_ls]).item()
                    final_col = col_ls[-1]
                    if final_col['max_length'] < avg_length:
                        final_col['max_length'] = round(avg_length, 2)
                    sheet_dict[name] = final_col

                self.sheet_dict = sheet_dict
                self.sheet_names = list(sheet_dict.keys())
                return sheet_dict
            else:
                if sheet_name not in sheet_ls:
                    raise KeyError(f"Sheet '{sheet_name}' not found in workbook.")
                source_ws = source_wb[sheet_name]
                col_ls = self.get_max_length_all_cols(source_ws)
                return {sheet_name: col_ls}
        finally:
            source_wb.close()

    def autofit_columns_spire(self, input_file: str, output_dir: str) -> str:
        from spire.xls import Workbook
        """Adjust last-column widths in Excel sheets using Spire.XLS.

        Uses ``self.sheet_names`` and ``self.sheet_dict`` populated by
        ``get_last_column_autofit``.

        Args:
            input_file (str): Path to the input Excel file.
            output_dir (str): Directory for the output Excel file.

        Returns:
            str: Path to the exported Excel file.
        """
        workbook = Workbook()
        try:
            workbook.LoadFromFile(input_file)
            for sheet_name in self.sheet_names:
                sheet = workbook.Worksheets.get_Item(sheet_name)
                if sheet is None:
                    logger.warning(f"Sheet '{sheet_name}' not found in Spire workbook. Skipping.")
                    continue
                adj = self.sheet_dict.get(sheet_name, {})
                last_col_idx = int(adj['index']) + 1
                adj_width = float(adj['max_length'])
                sheet.SetColumnWidth(last_col_idx, adj_width)

            output_file = os.path.join(output_dir, f"processed_{Path(input_file).stem}.xlsx")
            workbook.SaveToFile(output_file)
            logger.info(f"Processed Excel file exported: {output_file}")
            return str(output_file)
        finally:
            workbook.Dispose()

    def fix_diagram_per_sheet(self, process_excel_path: str, excel_process_dir: str) -> Path:
        from spire.xls import Workbook
        """Fix and resize diagrams/images in Excel sheets, then save the modified workbook.

        Resizes images larger than 100×100 px to 90 % of their original dimensions
        and shifts their row anchors. Uses ``self.sheet_names`` populated by
        ``get_last_column_autofit``. Result is saved with a ``fixed_layout_`` prefix.

        Args:
            process_excel_path (str): Path to the Excel file to process.
            excel_process_dir (str): Directory where the fixed file will be saved.

        Returns:
            Path: Path to the saved fixed-layout Excel file.

        Raises:
            Exception: If loading, processing, or saving the workbook fails.
        """
        workbook = Workbook()
        try:
            workbook.LoadFromFile(str(process_excel_path))

            for sheet_name in self.sheet_names:
                sheet = workbook.Worksheets[sheet_name]
                for i in range(len(sheet.Pictures) - 1, -1, -1):
                    pic = sheet.Pictures[i]
                    if pic.Width > 100 and pic.Height > 100:
                        pic.Width = int(pic.Width * 0.9)
                        pic.Height = int(pic.Height * 0.9)
                        pic.TopRow += 6
                        pic.BottomRow += 5

            fix_layout_excel_path = (
                Path(excel_process_dir) / f"fixed_layout_{Path(process_excel_path).stem}.xlsx"
            )
            workbook.CalculateAllValue()
            workbook.SaveToFile(str(fix_layout_excel_path))
            logger.info(f"Fixed diagrams and saved to: {fix_layout_excel_path}")
            return fix_layout_excel_path

        except Exception as e:
            raise Exception(f"Error processing diagrams in Excel file: {e}") from e
        finally:
            workbook.Dispose()

    # -------- METHOD 1: PREPROCCESSING EXCEL WITH KEEPING ALL IMAGES AND SHAPES  --------
    def preprocess_1st_method(self, file_bytes: bytes) -> bytes:
        """Preprocess Excel bytes before pipeline conversion.

        Writes the bytes to a temporary directory, runs the three preprocessing
        steps, and returns the final processed bytes.

        Steps:
            1. Compute per-sheet last-column autofit widths via openpyxl.
            2. Apply column-width adjustments and export via Spire.XLS.
            3. Fix/resize embedded diagrams and export the final file.

        Args:
            file_bytes (bytes): Raw Excel file content.

        Returns:
            bytes: Preprocessed Excel file content.
        """
        with tempfile.TemporaryDirectory() as _tmp_dir:
            tmp_dir = Path(_tmp_dir)
            input_path = tmp_dir / "input.xlsx"
            input_path.write_bytes(file_bytes)

            excel_process_dir = tmp_dir / 'excel_temp'
            excel_process_dir.mkdir()

            # Step 1: Compute per-sheet last-column autofit widths
            # Populates self.sheet_dict and self.sheet_names
            self.get_last_column_autofit(excel_path=str(input_path), get_all_sheets=True)

            # Step 2: Adjust column widths (uses self.sheet_names, self.sheet_dict)
            process_excel_path = self.autofit_columns_spire(
                input_file=str(input_path),
                output_dir=str(excel_process_dir),
            )

            # Step 3: Fix/resize embedded diagrams (uses self.sheet_names)
            fix_layout_excel_path = self.fix_diagram_per_sheet(
                process_excel_path=str(process_excel_path),
                excel_process_dir=str(excel_process_dir),
            )

            return Path(fix_layout_excel_path).read_bytes()
    

    # -------- METHOD 2: ENHANCED EXCEL FILE BY REMOVING IMG, OLE OBJECTS, EMPTY COLS, ROWS --------
    def preprocess_2nd_method(self, file_bytes: bytes) -> bytes:
        """Preprocess Excel bytes before pipeline conversion.

        Writes the bytes to a temporary directory, runs the three preprocessing
        steps, and returns the final processed bytes.

        Steps:
            1. Compute per-sheet last-column autofit widths via openpyxl.
            2. Apply column-width adjustments and export via Spire.XLS.
            3. Fix/resize embedded diagrams and export the final file.

        Args:
            file_bytes (bytes): Raw Excel file content.

        Returns:
            bytes: Preprocessed Excel file content.
        """
        with tempfile.TemporaryDirectory() as _tmp_dir:
            tmp_dir = Path(_tmp_dir)
            input_path = tmp_dir / "input.xlsx"
            input_path.write_bytes(file_bytes)

            excel_process_dir = tmp_dir / 'excel_temp'
            excel_process_dir.mkdir()

            # Step 1: clean excel file: remove all shapes, imgs and ole objects 
            #self.get_last_column_autofit(excel_path=str(input_path), get_all_sheets=True)
            optimize_excel_path = clean_and_optimize_excel(excel_path=str(input_path),output_dir=str(excel_process_dir))

            # Step 2: analyse and adjust excel file by Sprire
            final_adjust_path, adjust_data = analyze_and_adjust_excel_dimensions(file_path=optimize_excel_path, 
                                                                                output_dir=str(Path(optimize_excel_path).parent))

            print(f"Final adjust path: {final_adjust_path}")
            return Path(final_adjust_path).read_bytes()

    # COMBINE BOTH PROCESS: COPY INVALID SHEET NAME FROM METHOD 2 INTO METHOD 1
    # Checks the data range of all sheets in a workbook using Spire.XLS.
    @staticmethod
    def validate_sheet_sizes(file_path: str, max_row_limit: int = 100, max_col_limit: int = 100) -> dict:
        """
        Checks the data range of all sheets in a workbook using Spire.XLS.
        Returns a dictionary categorizing sheets as valid or invalid based on size.
        """
        from spire.xls import Workbook
        workbook = Workbook()
        workbook.LoadFromFile(file_path)
        
        validation_status: dict[str, list] = {
            "valid_sheets": [],
            "invalid_sheets": []
        }
        
        for sheet in workbook.Worksheets:
            sheet_name = sheet.Name
            data_range = sheet.AllocatedRange
            
            # If the sheet is completely empty, it is technically not oversize
            if data_range is None:
                validation_status["valid_sheets"].append(sheet_name)
                continue
                
            last_row = data_range.LastRow
            last_col = data_range.LastColumn
            
            # Check against your defined thresholds
            if last_row > max_row_limit or last_col > max_col_limit:
                validation_status["invalid_sheets"].append({
                    "sheet_name": sheet_name,
                    "last_row": last_row,
                    "last_col": last_col
                })
                print(f"Warning: Sheet '{sheet_name}' is invalid (Oversize: {last_row} rows, {last_col} cols).")
            else:
                validation_status["valid_sheets"].append(sheet_name)
                
        workbook.Dispose()
        return validation_status
    
    # Copies specific sheets from source to destination. 
    @staticmethod
    def copy_specific_sheets(source_file: str, dest_file: str, new_output_file: str, sheets_to_copy: list):
        """
        Copies specific sheets from source to destination. 
        Safely handles name collisions and Excel's 31-character limit.
        """
        from spire.xls import Workbook
        source_wb = Workbook()
        source_wb.LoadFromFile(source_file)
        
        dest_wb = Workbook()
        dest_wb.LoadFromFile(dest_file)
        
        for sheet_name in sheets_to_copy:
            source_sheet = source_wb.Worksheets[sheet_name]
            
            if source_sheet is not None:
                # 1. Get a list of all current sheet names in the destination
                existing_names = [sheet.Name for sheet in dest_wb.Worksheets]
                target_name = sheet_name
                
                # 2. Generate a guaranteed unique name that respects the 31-char limit
                if target_name in existing_names:
                    # Truncate the base name to 26 chars so "_copy" (5 chars) fits the 31-char limit
                    base_name = sheet_name[:26] + "_copy"
                    target_name = base_name
                    counter = 1
                    
                    # If the _copy name also exists, append a number
                    while target_name in existing_names:
                        suffix = f"{counter}"
                        # Ensure base + suffix never exceeds 31 characters
                        target_name = base_name[:31 - len(suffix)] + suffix
                        counter += 1
                
                # 3. Perform the copy
                new_sheet = dest_wb.Worksheets.AddCopy(source_sheet)
                
                # 4. Safely rename ONLY if Spire hasn't already assigned the target name
                if new_sheet.Name != target_name:
                    new_sheet.Name = target_name
                
                print(f"Successfully copied sheet as: '{target_name}'")
            else:
                print(f"Warning: Sheet '{sheet_name}' not found in source file. Skipping.")
                
        # Save and clean up
        dest_wb.SaveToFile(new_output_file)
        source_wb.Dispose()
        dest_wb.Dispose()
        print(f"Process complete. New file saved as: {new_output_file}")
 
    # Preprocess Excel bytes before pipeline conversion.
    def preprocess(self, file_bytes: bytes) -> bytes:
        """Preprocess Excel bytes before pipeline conversion.

        Writes the bytes to a temporary directory, runs the three preprocessing
        steps, and returns the final processed bytes.

        Steps:
            1. Compute per-sheet last-column autofit widths via openpyxl.
            2. Apply column-width adjustments and export via Spire.XLS.
            3. Fix/resize embedded diagrams and export the final file.

        Args:
            file_bytes (bytes): Raw Excel file content.

        Returns:
            bytes: Preprocessed Excel file content.
        """
        with tempfile.TemporaryDirectory() as _tmp_dir:
            tmp_dir = Path(_tmp_dir)
            excel_process_dir = tmp_dir / 'excel_temp'
            excel_process_dir.mkdir()

            # apply method 1 
            output1_path = excel_process_dir / "output1.xlsx"
            process_bytes_m1 = self.preprocess_1st_method(file_bytes=file_bytes)
            output1_path.write_bytes(process_bytes_m1)

            # apply method 2
            output2_path = excel_process_dir / "output2.xlsx"
            process_bytes_m2 = self.preprocess_2nd_method(file_bytes=file_bytes)
            output2_path.write_bytes(process_bytes_m2)

            # Check validation for oversize sheeet, then get invalid sheets.
            validation_status = self.validate_sheet_sizes(file_path=str(output1_path.resolve()))
            sheet_ls = [i['sheet_name'] for i in validation_status['invalid_sheets']]

            # Generate the final excel file include method 1 and method 2 (a parts)
            if not sheet_ls:
                return Path(output1_path).read_bytes()
            final_output_path = excel_process_dir / "output_final.xlsx"
            try:
                self.copy_specific_sheets(source_file=str(output2_path.resolve()),
                                        dest_file=str(output1_path.resolve()),
                                        new_output_file=str(final_output_path.resolve()),
                                        sheets_to_copy=sheet_ls)
            except Exception as e: 
                logger.error(f"copy_specific_sheets failed: {e}", exc_info=True)
                raise

            return Path(final_output_path).read_bytes()

